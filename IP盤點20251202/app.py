import openpyxl
import csv
import os
from datetime import datetime

def determine_category(ip_address, machine_id, device_name):
    """根據 IP 地址、Machine ID 和 Device Name 判斷 Category
    
    規則:
    1. 如果 IP 最後一段 <= 40,Category = '\xa0EAP' (非中斷空格 + EAP)
    2. 如果 IP 最後一段 >= 231,Category = 'Switch'
    3. 如果包含 "Switch" 關鍵字,Category = 'Switch'
    4. 否則 Category = 'EQP'
    """
    # 轉換為字串並處理 None
    machine_id_str = str(machine_id).strip() if machine_id else ""
    device_name_str = str(device_name).strip() if device_name else ""
    
    # 提取 IP 最後一段
    if ip_address and ip_address != 'None':
        ip_str = str(ip_address).strip()
        if '.' in ip_str:
            try:
                ip_last_octet = int(ip_str.split('.')[-1])
                # 如果 IP 尾數 <= 40,視為 EAP
                if ip_last_octet <= 40:
                    return '\xa0EAP'  # 使用非中斷空格 \xa0
                # 如果 IP 尾數 >= 231,視為 Switch
                elif ip_last_octet >= 231:
                    return 'Switch'
            except (ValueError, IndexError):
                pass
    
    # 檢查是否為 Switch (關鍵字判斷)
    combined = (machine_id_str + " " + device_name_str).upper()
    if "SWITCH" in combined:
        return "Switch"
    
    # 預設為 EQP
    return "EQP"

def determine_alive_status(row_data):
    """判斷設備是否 alive 或 dead"""
    # 如果 Internal IP、Machine ID 或 Device Name 有值,通常視為 alive
    internal_ip = row_data.get('Internal_IP', '')
    machine_id = row_data.get('Machine_ID', '')
    device_name = row_data.get('Device_Name', '')
    
    # 如果 IP 是 0 或沒有任何識別資訊,視為 dead
    if internal_ip == '0' or (not machine_id and not device_name):
        return "dead"
    
    # 如果有 Device Name,視為 alive
    if device_name and device_name != 'None':
        return "alive"
    
    # 預設為 dead
    return "dead"

def format_date(date_value):
    """格式化日期 - Windows 兼容版本"""
    if isinstance(date_value, datetime):
        # 使用 int() 去除前導零
        year = date_value.year
        month = date_value.month
        day = date_value.day
        return f"{year}/{month}/{day}"
    elif date_value:
        return str(date_value)
    return ''

def parse_sheet_info(sheet_name):
    """從工作表名稱解析棟別、樓層和完整位置資訊
    
    例如: 'K11-3F 區網(27)' -> 
        building='K11', floor='3F', location='K11-3F', file_place='K11\\3F\\K11-3F 區網(27).csv'
    """
    # 例如: K11-3F 區網(27)
    parts = sheet_name.split(' ')
    if len(parts) > 0:
        location = parts[0]  # K11-3F
        building_floor = location.split('-')
        if len(building_floor) >= 2:
            building = building_floor[0]  # K11
            floor = building_floor[1]  # 3F
            file_place = f"{building}\\{floor}\\{sheet_name}.csv"
            return building, floor, location, file_place
    
    # 如果無法解析,回傳預設值
    return 'K11', '3F', sheet_name, f"{sheet_name}.csv"

def is_yellow_row(ws, row_idx):
    """檢查該行是否有黃色背景"""
    # 檢查該行的所有儲存格
    for cell in ws[row_idx]:
        if cell.fill and cell.fill.fgColor:
            # 取得顏色值
            color = cell.fill.fgColor
            if color.rgb:
                # 黃色的 RGB 值通常是 FFFFFF00 或 FFFF00
                color_hex = str(color.rgb).upper()
                # 檢查是否為黃色系 (RGB 中 R 和 G 較高, B 較低)
                if 'FFFF' in color_hex or 'FFF' in color_hex[:6]:
                    # 進一步確認是黃色 (不是白色)
                    if color_hex not in ['FFFFFFFF', 'FFFFFF']:
                        return True
    return False

def convert_xlsx_to_csv(xlsx_path, output_dir):
    """將 xlsx 檔案的每個工作表轉換為 CSV"""
    
    # 讀取 xlsx (read_only=False 才能讀取格式)
    wb = openpyxl.load_workbook(xlsx_path, read_only=False)
    
    # CSV 欄位名稱
    csv_columns = [
        'Internal_IP', 'Machine_ID', 'Local', 'Device_Name', 
        'TCP_Port', 'COM_Port', 'OS_Spec', 'IP_Source', 
        'Category', 'Online_Test', 'Set_Time', 'Remark', 
        '歲修', 'File_Place', '所在區域(柱位)', 'alive_or_dead'
    ]
    
    # 處理每個工作表
    for sheet_name in wb.sheetnames:
        # 跳過不需要的工作表
        if sheet_name == '工作表1':
            continue
            
        print(f"處理工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        # 解析棟別、樓層資訊
        building, floor, location, file_place = parse_sheet_info(sheet_name)
        
        # 創建輸出資料夾結構: output_dir/棟別/樓層/
        output_folder = os.path.join(output_dir, building, floor)
        os.makedirs(output_folder, exist_ok=True)
        
        # 創建輸出 CSV 檔案路徑
        output_filename = f"{sheet_name}.csv"
        output_path = os.path.join(output_folder, output_filename)
        
        # 寫入 CSV
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
            
            # 寫入標題行
            writer.writeheader()
            
            # 從第4行開始讀取資料 (跳過標題和說明行)
            for row_idx in range(4, ws.max_row + 1):
                # 檢查是否為黃色標記行
                if is_yellow_row(ws, row_idx):
                    # 寫入全 0 的行
                    zero_row = {
                        'Internal_IP': '0',
                        'Machine_ID': '0',
                        'Local': '0',
                        'Device_Name': '0',
                        'TCP_Port': '0',
                        'COM_Port': '0',
                        'OS_Spec': '0',
                        'IP_Source': '0',
                        'Category': '0',
                        'Online_Test': '0',
                        'Set_Time': '0',
                        'Remark': '0',
                        '歲修': '0',
                        'File_Place': '0',
                        '所在區域(柱位)': '',
                        'alive_or_dead': 'dead'
                    }
                    writer.writerow(zero_row)
                    continue
                
                # 正常處理非黃色行
                row = [cell.value for cell in ws[row_idx]]
                
                # XLSX 的欄位對應 (第一列是空的)
                internal_ip = row[1] if len(row) > 1 else None
                machine_id = row[2] if len(row) > 2 else None
                # local = row[3]  # 不再使用原始的 local,改用統一的 location
                device_name = row[4] if len(row) > 4 else None
                tcp_port = row[5] if len(row) > 5 else None
                com_port = row[6] if len(row) > 6 else None
                os_spec = row[7] if len(row) > 7 else None
                ip_source = row[8] if len(row) > 8 else None
                # row[9] 是 New Rule,不使用
                online_test = row[10] if len(row) > 10 else None
                set_time = row[11] if len(row) > 11 else None
                remark = row[12] if len(row) > 12 else None
                
                # 跳過完全空白的行
                if not any([internal_ip, machine_id, device_name]):
                    continue
                
                # 格式化資料
                internal_ip_str = str(internal_ip).strip() if internal_ip and internal_ip != 'None' else ''
                machine_id_str = str(machine_id).strip() if machine_id and machine_id != 'None' else ''
                device_name_str = str(device_name).strip() if device_name and device_name != 'None' else ''
                tcp_port_str = str(tcp_port).strip() if tcp_port and tcp_port != 'None' else ''
                com_port_str = str(com_port).strip() if com_port and com_port != 'None' else ''
                os_spec_str = str(os_spec).strip() if os_spec and os_spec != 'None' else ''
                ip_source_str = str(ip_source).strip() if ip_source and ip_source != 'None' else ''
                online_test_str = str(online_test).strip() if online_test and online_test != 'None' else ''
                set_time_str = format_date(set_time)
                remark_str = str(remark).strip() if remark and remark != 'None' else ''
                
                # 建立行資料字典
                row_data = {
                    'Internal_IP': internal_ip_str,
                    'Machine_ID': machine_id_str,
                    'Device_Name': device_name_str
                }
                
                # 判斷 Category
                category = determine_category(internal_ip_str, machine_id, device_name)
                
                # 判斷 alive_or_dead
                alive_status = determine_alive_status(row_data)
                
                # 組合完整的行資料
                csv_row = {
                    'Internal_IP': internal_ip_str,
                    'Machine_ID': machine_id_str,
                    'Local': location,  # 統一使用解析出的 location (例如: K11-3F)
                    'Device_Name': device_name_str,
                    'TCP_Port': tcp_port_str,
                    'COM_Port': com_port_str,
                    'OS_Spec': os_spec_str,
                    'IP_Source': ip_source_str,
                    'Category': category,
                    'Online_Test': online_test_str,
                    'Set_Time': set_time_str,
                    'Remark': remark_str,
                    '歲修': '',  # 預設為空
                    'File_Place': file_place,  # 使用解析出的 file_place
                    '所在區域(柱位)': '',  # 預設為空
                    'alive_or_dead': alive_status
                }
                
                writer.writerow(csv_row)
        
        print(f"  已生成: {building}/{floor}/{output_filename}")
    
    wb.close()
    print(f"\n轉換完成!")

# 執行轉換
if __name__ == '__main__':
    xlsx_file = 'K25_New_EQP_IP_List_20251202_(Security C).xlsx'
    output_directory = './output'
    
    # 確保輸出目錄存在
    os.makedirs(output_directory, exist_ok=True)
    
    convert_xlsx_to_csv(xlsx_file, output_directory)